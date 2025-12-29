import os
import pandas as pd
from openpyxl.styles import Alignment

# 시나리오 리스트 txt 파일 경로
scenario_txt_path = "/home/ubuntu/Dataset/000.datasetStructure/01.data_acquisitionStructure/scenario_list.txt"
# 기준 디렉토리 (시나리오 폴더들이 들어있는 상위 폴더)
base_dir = "../../Dataset/08_BLTN_GEN3/99_SCENARIO_SET_IMAGE_448x224"
output_excel_path = "scenario_checklist_251124.xlsx" # 출력 엑셀 파일
depth = 1

# 시나리오 리스트 로드
with open(scenario_txt_path, 'r') as f:
    scenario_list = [line.strip() for line in f if line.strip()]

# 2-depth 모든 디렉토리 경로 수집
all_dirs = set()
if depth == 2:
    for root, dirs, files in os.walk(base_dir):
        for d in dirs:
            full_path = os.path.join(root, d)
            all_dirs.add(os.path.basename(full_path)) # 폴더 이름만 저장
            print(full_path)
elif depth == 1:
    for item in os.listdir(base_dir):
        item_path = os.path.join(base_dir, item)
        if os.path.isdir(item_path):
            all_dirs.add(item)
            print(item_path)

# 존재 여부 판별
data = []
for scenario in scenario_list:
    acquired = 0
    scenario_path = os.path.join(base_dir, scenario)
    if os.path.isdir(scenario_path):
        try:
            has_file = any(
                os.path.isfile(os.path.join(scenario_path, f))
                for f in os.listdir(scenario_path)
            )
            if has_file:
                acquired = 1
        except Exception as e:
            print(f"[오류] {scenario_path} 접근 실패: {e}")
            acquired = 0
    
    splitted = scenario.split('_')
    splitted.append(acquired)
    data.append(splitted)

# 최대 컬럼 수 계산
max_columns = max(len(row) for row in data)
column_names = ['환경', '투과율', '카메라방향', '객체', '오염물질', '취득여부']

# 각 행을 오른쪽 정렬 (왼쪽에 빈칸 추가)
aligned_data = []
for row in data:
    # 부족한 만큼 왼쪽에 None(빈칸) 추가
    empty_cells = max_columns - len(row)
    aligned_row = [None] * empty_cells + row
    aligned_data.append(aligned_row)

# 컬럼명도 오른쪽 정렬
aligned_columns = [None] * (max_columns - len(column_names)) + column_names

# DataFrame 생성 및 저장
df = pd.DataFrame(aligned_data, columns=aligned_columns)

# ExcelWriter로 저장
with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Sheet1")
    ws = writer.sheets["Sheet1"]
    
    # 텍스트 정렬 (선택사항)
    center_align = Alignment(horizontal="center", vertical="center")
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_align

print(f"[완료] 엑셀 파일 저장됨: {output_excel_path}")
